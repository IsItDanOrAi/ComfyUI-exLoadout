import openpyxl
import os
from folder_paths import get_filename_list, get_full_path_or_raise, get_folder_paths
import comfy.sd

def get_excel_full_path_or_raise(base_folder, file_path):
    """
    Securely resolve Excel file paths within a designated directory.
    
    Args:
        base_folder: The base folder name (use "." for current directory)
        file_path: The requested file path
        
    Returns:
        str: The absolute path if valid
        
    Raises:
        ValueError: If the path is invalid or outside the allowed directory
    """
    # Get the directory where the script is located
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # If base_folder is ".", use the current directory, otherwise create subdirectory path
    if base_folder == ".":
        base_dir = current_dir
    else:
        base_dir = os.path.join(current_dir, base_folder)
    
    # Normalize the file path to prevent directory traversal
    normalized_file_path = os.path.normpath(file_path)
    
    # Check for directory traversal attempts
    if os.path.isabs(normalized_file_path) or normalized_file_path.startswith('..'):
        raise ValueError("Invalid file path. Absolute paths and parent directory references are not allowed.")
    
    # Construct the full path
    full_path = os.path.join(base_dir, normalized_file_path)
    
    # Resolve any remaining relative components
    resolved_path = os.path.abspath(full_path)
    
    # Ensure the resolved path is still within the base directory
    if not resolved_path.startswith(os.path.abspath(base_dir)):
        raise ValueError("Invalid file path. Path must be within the designated directory.")
    
    return resolved_path

class exLoadoutCheckpointLoader:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "exLoadoutList.xlsx"}),  # Just the filename
                "sheet_name": ("STRING", {"default": "MODELS"}),
                "loadout_name": ("STRING", {"default": ""}),
                "clip_type": (["stable_diffusion", "stable_cascade", "sd3", "stable_audio", "mochi", "ltxv", "pixart", "cosmos", "lumina2", "wan"],),
            },
        }

    RETURN_TYPES = ("MODEL", "CLIP", "VAE", "STRING")
    RETURN_NAMES = ("model", "clip", "vae", "Output")
    FUNCTION = "exLoadoutCheckpointLoader"
    CATEGORY = "exLoadout"
    DESCRIPTION = (
        "Loads a checkpoint model by reading its name from Column B, "
        "CLIP from Column C, and VAE from Column D in an Excel file. "
        "Each row is identified by a 'Loadout' name from Column A."
    )

    def exLoadoutCheckpointLoader(self, excel_path, sheet_name, loadout_name, clip_type):
        # Secure path resolution for Excel file - look in current directory
        full_excel_path = get_excel_full_path_or_raise(".", excel_path)

        # Validate file extension
        if not full_excel_path.lower().endswith(".xlsx"):
            raise ValueError("Invalid file type. Only .xlsx files are supported.")

        # Check if file exists
        if not os.path.exists(full_excel_path):
            base_dir = os.path.dirname(os.path.abspath(__file__))
            raise FileNotFoundError(f"Excel file not found: {os.path.basename(full_excel_path)}\n"
                                  f"Expected location: {full_excel_path}\n"
                                  f"Make sure the file exists in: {base_dir}")

        workbook = openpyxl.load_workbook(full_excel_path)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")

        sheet = workbook[sheet_name]
        found_row = None

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            if row[0].value and str(row[0].value).strip() == loadout_name:
                found_row = row
                break

        workbook.close()

        if not found_row:
            raise ValueError(f"Loadout '{loadout_name}' not found in Column A.")

        # Load checkpoint model (Column B)
        if not found_row[1].value:
            raise ValueError(f"No valid checkpoint name found for Loadout '{loadout_name}' in Column B.")
        ckpt_name = str(found_row[1].value).strip()

        allowed_ckpts = get_filename_list("checkpoints")
        if ckpt_name not in allowed_ckpts:
            raise ValueError(f"Checkpoint '{ckpt_name}' is not in the allowed checkpoints list.")

        # Use ComfyUI's secure path resolution for model files
        ckpt_path = get_full_path_or_raise("checkpoints", ckpt_name)
        model, clip, vae = comfy.sd.load_checkpoint_guess_config(
            ckpt_path,
            output_vae=True,
            output_clip=True,
            embedding_directory=get_folder_paths("embeddings")
        )[:3]

        # Load CLIP (Column C)
        clip_name = "Default"
        if len(found_row) > 2 and found_row[2].value:
            temp_clip_name = str(found_row[2].value).strip()
            if temp_clip_name in get_filename_list("text_encoders"):
                try:
                    # Use ComfyUI's secure path resolution for CLIP files
                    clip_path = get_full_path_or_raise("text_encoders", temp_clip_name)
                    clip = comfy.sd.load_clip(
                        ckpt_paths=[clip_path],
                        embedding_directory=get_folder_paths("embeddings"),
                        clip_type=clip_type
                    )
                    clip_name = temp_clip_name
                except Exception as e:
                    print(f"Warning: Failed to load CLIP override '{temp_clip_name}': {e}")

        # Load VAE (Column D)
        vae_name = "Default"
        if len(found_row) > 3 and found_row[3].value:
            temp_vae_name = str(found_row[3].value).strip()
            if temp_vae_name in get_filename_list("vae"):
                try:
                    # Use ComfyUI's secure path resolution for VAE files
                    vae_path = get_full_path_or_raise("vae", temp_vae_name)
                    vae = comfy.sd.load_vae(vae_path)
                    vae_name = temp_vae_name
                except Exception as e:
                    print(f"Warning: Failed to load VAE override '{temp_vae_name}': {e}")

        debug_output = f"Loadout: {loadout_name}, Model: {ckpt_name}, CLIP: {clip_name}, VAE: {vae_name}"
        return (model, clip, vae, debug_output)

NODE_CLASS_MAPPINGS = {"exLoadoutCheckpointLoader": exLoadoutCheckpointLoader}
NODE_DISPLAY_NAME_MAPPINGS = {"exLoadoutCheckpointLoader": "exLoadout Checkpoint Loader"}
