import openpyxl
import os
import random
import time
import comfy.sd

def get_excel_full_path_or_raise(base_folder, file_path):
    """
    Securely resolve Excel file paths within a designated directory.
    
    Args:
        base_folder: The base folder name (use "." for current directory)
        file_path: The requested file path
        
    Returns:
        str: The absolute path if valid
        
    Raises:
        ValueError: If the path is invalid or outside the allowed directory
    """
    # Get the directory where the script is located
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # If base_folder is ".", use the current directory, otherwise create subdirectory path
    if base_folder == ".":
        base_dir = current_dir
    else:
        base_dir = os.path.join(current_dir, base_folder)
    
    # Normalize the file path to prevent directory traversal
    normalized_file_path = os.path.normpath(file_path)
    
    # Check for directory traversal attempts
    if os.path.isabs(normalized_file_path) or normalized_file_path.startswith('..'):
        raise ValueError("Invalid file path. Absolute paths and parent directory references are not allowed.")
    
    # Construct the full path
    full_path = os.path.join(base_dir, normalized_file_path)
    
    # Resolve any remaining relative components
    resolved_path = os.path.abspath(full_path)
    
    # Ensure the resolved path is still within the base directory
    if not resolved_path.startswith(os.path.abspath(base_dir)):
        raise ValueError("Invalid file path. Path must be within the designated directory.")
    
    return resolved_path

class exLoadoutSelector:
    # Class variable to track the current index for sequential selection
    _current_index = 0
    
    @classmethod
    def NODE_NAME(cls):
        """Sets the node name to 'exLoadout Selector' instead of the class name."""
        return "exLoadout Selector"
    
    @classmethod
    def INPUT_TYPES(cls):
        # Get dynamic options from Excel file
        excel_path = "exLoadoutList.xlsx"
        sheet_name = "MODELS"
        dynamic_options, default_value = cls.get_excel_options(excel_path, sheet_name)
        
        return {
            "required": {
                "excel_path": ("STRING", {"default": excel_path}),
                "sheet_name": ("STRING", {"default": sheet_name}),
                "Loadout": (dynamic_options, {"default": default_value}),  # Dynamic options from Excel
                "selection_mode": (["Random", "Increment", "Decrement"], {"default": "Random"}),  # Selection mode
            },
        }
    
    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("Loadout", "Auto Loadout")
    FUNCTION = "get_selected_loadout"
    CATEGORY = "exLoadout"
    DESCRIPTION = "Dropdown populated from Column A of an Excel file. Returns the selected Loadout and an auto-selected loadout based on mode."
    
    @classmethod
    def IS_CHANGED(cls, excel_path, sheet_name, Loadout, selection_mode):
        """This method tells ComfyUI when to refresh the node's options."""
        try:
            full_excel_path = get_excel_full_path_or_raise(".", excel_path)
            if os.path.exists(full_excel_path):
                # Return a combination of file modification time and current time to force refresh
                return str(os.path.getmtime(full_excel_path)) + str(time.time())
        except:
            pass
        return str(time.time())  # Always refresh with current time if there's an issue
    
    @classmethod
    def get_excel_data(cls, excel_path, sheet_name):
        """Reads Column A from the Excel file and returns both all options and non-empty options."""
        try:
            # Secure path resolution for Excel file - look in current directory
            full_excel_path = get_excel_full_path_or_raise(".", excel_path)
            
            # Validate file extension
            if not full_excel_path.lower().endswith(".xlsx"):
                print("Error: Invalid file type. Only .xlsx files are supported.")
                return ["ERROR: INVALID FILE TYPE"], "ERROR: INVALID FILE TYPE", []
            
            # Check if file exists
            if not os.path.exists(full_excel_path):
                base_dir = os.path.dirname(os.path.abspath(__file__))
                print(f"Excel file not found: {os.path.basename(full_excel_path)}")
                print(f"Expected location: {full_excel_path}")
                print(f"Make sure the file exists in: {base_dir}")
                return ["ERROR: FILE NOT FOUND"], "ERROR: FILE NOT FOUND", []
                
        except Exception as e:
            print(f"Path error: {e}")
            return ["ERROR: FILE NOT FOUND"], "ERROR: FILE NOT FOUND", []
        
        try:
            workbook = openpyxl.load_workbook(full_excel_path, read_only=True)
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                print(f"Error: Sheet '{sheet_name}' not found in Excel file.")
                return ["ERROR: SHEET NOT FOUND"], "ERROR: SHEET NOT FOUND", []
            
            sheet = workbook[sheet_name]
            
            # Read Column A starting from row 2 (skip header row A1)
            options = []
            non_empty_options = []
            first_value = None
            
            for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 to skip header
                cell_value = sheet.cell(row=row_idx, column=1).value
                if cell_value is not None:
                    value = str(cell_value).strip()
                    if value:  # Only add non-empty strings
                        options.append(value)
                        non_empty_options.append(value)
                        if first_value is None:  # Store the first non-empty value (A2)
                            first_value = value
                else:
                    options.append("empty")
                    if first_value is None:  # If A2 is empty, set first_value to "empty"
                        first_value = "empty"
            
            workbook.close()
            
            # If no data found, return empty option
            if not options:
                options = ["empty"]
                first_value = "empty"
            
            # Debug print to help troubleshoot
            print(f"Excel options found: {options}")
            print(f"Non-empty options: {non_empty_options}")
            print(f"Default value (A2): {first_value}")
            
            return options, first_value, non_empty_options
        
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return ["ERROR: READ FAILED"], "ERROR: READ FAILED", []
    
    @classmethod
    def get_excel_options(cls, excel_path, sheet_name):
        """Reads Column A from the Excel file and returns a list of options and default value."""
        options, first_value, _ = cls.get_excel_data(excel_path, sheet_name)
        return options, first_value
    
    def get_selected_loadout(self, excel_path, sheet_name, Loadout, selection_mode):
        """Returns the selected Loadout value from Column A and an auto-selected loadout based on mode."""
        # Get all data in one call to avoid multiple file reads
        options, _, non_empty_options = self.get_excel_data(excel_path, sheet_name)
        
        # Handle selected loadout
        if Loadout not in options or Loadout.startswith("ERROR:"):
            selected_loadout = "empty"
        else:
            selected_loadout = Loadout
        
        # Get auto loadout based on selection mode
        if not non_empty_options:
            auto_loadout = "Sheet is blank"
        else:
            if selection_mode == "Random":
                auto_loadout = random.choice(non_empty_options)
                print(f"Random selection from {non_empty_options}: {auto_loadout}")
            elif selection_mode == "Increment":
                # Use class variable to track current position for sequential selection
                auto_loadout = non_empty_options[self.__class__._current_index % len(non_empty_options)]
                print(f"Increment selection from {non_empty_options}: {auto_loadout} (index: {self.__class__._current_index})")
                self.__class__._current_index += 1
            elif selection_mode == "Decrement":
                # Use class variable to track current position for reverse sequential selection
                auto_loadout = non_empty_options[-(self.__class__._current_index % len(non_empty_options)) - 1]
                print(f"Decrement selection from {non_empty_options}: {auto_loadout} (index: {self.__class__._current_index})")
                self.__class__._current_index += 1
            else:
                # Fallback to random if mode is unrecognized
                auto_loadout = random.choice(non_empty_options)
                print(f"Fallback random selection: {auto_loadout}")
        
        return (selected_loadout, auto_loadout)

NODE_CLASS_MAPPINGS = {"exLoadoutSelector": exLoadoutSelector}
NODE_DISPLAY_NAME_MAPPINGS = {"exLoadoutSelector": "exLoadout Selector"}
