import openpyxl
import os
from typing import Union

# Hack: string type that is always equal in not equal comparisons
class AnyType(str):
    def __ne__(self, __value: object) -> bool:
        return False

# Our any instance wants to be a wildcard string
ANY = AnyType("*")

class exLoadoutSeg:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": os.path.join("exLoadoutList.xlsx")}),
                "sheet_name": ("STRING", {"default": "KSAMPLER"}),
                "row_number": ("INT", {"default": 1, "min": 1, "max": 1000, "step": 1}),
                "search_string": ("STRING", {"default": ""}),
            }
        }
    
    # Updated return types to include Column F
    RETURN_TYPES = (ANY, ANY, ANY, ANY, ANY, ANY, "STRING")
    RETURN_NAMES = ("Column A", "Column B", "Column C", "Column D", "Column E", "Column F", "Outputs")
    
    # Set all outputs except the summary to be lists
    OUTPUT_IS_LIST = (True, True, True, True, True, True, False)
    
    FUNCTION = "process_excel"
    CATEGORY = "exLoadout"
    DESCRIPTION = ("Reads values from columns A through F for a specified row number in an Excel spreadsheet. Can also search for a string in Column A.")
    NAME = "exLoadoutSeg (List)"
    
    def process_excel(self, excel_path, sheet_name, row_number, search_string):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        full_excel_path = os.path.join(current_dir, excel_path)
        if not os.path.exists(full_excel_path):
            raise FileNotFoundError(f"Excel file not found: {full_excel_path}")
        
        workbook = openpyxl.load_workbook(full_excel_path)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")
        
        sheet = workbook[sheet_name]
        
        # If search_string is provided, look for it in Column A
        actual_row = row_number
        if search_string:
            found = False
            for row_idx in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=1).value
                if cell_value is not None and str(cell_value).strip() == search_string:
                    actual_row = row_idx
                    found = True
                    break
            
            if not found:
                workbook.close()
                raise ValueError(f"Search string '{search_string}' not found in Column A.")
        
        # Check if row number is valid
        if actual_row < 1 or actual_row > sheet.max_row:
            workbook.close()
            raise ValueError(f"Row number {actual_row} is out of range. The sheet has {sheet.max_row} rows.")
        
        # Get the row data for columns A through F, preserving original types
        row_data = [sheet.cell(row=actual_row, column=col).value 
                    for col in range(1, 7)]  # Columns A through F
        
        # Handle None values by converting to empty string
        row_data = ['' if value is None else value for value in row_data]
        
        workbook.close()
        
        # Create outputs summary with the requested format including % symbols
        # Convert to string for summary, but preserve original type in the lists
        outputs_summary = f"%A: {row_data[0]} %B: {row_data[1]} %C: {row_data[2]} %D: {row_data[3]} %E: {row_data[4]} %F: {row_data[5]} %"
        
        # Wrap each value in a list to maintain list compatibility
        return ([row_data[0]], [row_data[1]], [row_data[2]], [row_data[3]], [row_data[4]], [row_data[5]], outputs_summary)