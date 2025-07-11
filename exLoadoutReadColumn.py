import openpyxl
import os

class AnyType(str):
    def __ne__(self, __value: object) -> bool:
        return False

ANY = AnyType("*")

def get_excel_full_path_or_raise(base_folder, file_path):
    """
    Securely resolve Excel file paths within a designated directory.
    
    Args:
        base_folder: The base folder name (use "." for current directory)
        file_path: The requested file path
        
    Returns:
        str: The absolute path if valid
        
    Raises:
        ValueError: If the path is invalid or outside the allowed directory
    """
    # Get the directory where the script is located
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # If base_folder is ".", use the current directory, otherwise create subdirectory path
    if base_folder == ".":
        base_dir = current_dir
    else:
        base_dir = os.path.join(current_dir, base_folder)
    
    # Normalize the file path to prevent directory traversal
    normalized_file_path = os.path.normpath(file_path)
    
    # Check for directory traversal attempts
    if os.path.isabs(normalized_file_path) or normalized_file_path.startswith('..'):
        raise ValueError("Invalid file path. Absolute paths and parent directory references are not allowed.")
    
    # Construct the full path
    full_path = os.path.join(base_dir, normalized_file_path)
    
    # Resolve any remaining relative components
    resolved_path = os.path.abspath(full_path)
    
    # Ensure the resolved path is still within the base directory
    if not resolved_path.startswith(os.path.abspath(base_dir)):
        raise ValueError("Invalid file path. Path must be within the designated directory.")
    
    return resolved_path

class exLoadoutReadColumn:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "exLoadoutList.xlsx"}),  # Default Excel filename
                "sheet_name": ("STRING", {"default": "MODELS"}),  # Default sheet name
                "column_letter": ("STRING", {"default": "A"}),  # Column selection by letter
            },
        }
    
    RETURN_TYPES = (ANY,)
    RETURN_NAMES = ("output_list",)
    OUTPUT_IS_LIST = (True,)
    FUNCTION = "read_excel_column"
    CATEGORY = "exLoadout"
    DESCRIPTION = "Reads all values from a specified column in an Excel spreadsheet and returns them as a comma-separated string inside a list."
    
    def read_excel_column(self, excel_path, sheet_name, column_letter):
        # ✅ Secure path resolution for Excel file - look in current directory
        full_excel_path = get_excel_full_path_or_raise(".", excel_path)
        
        # Validate file extension
        if not full_excel_path.lower().endswith(".xlsx"):
            raise ValueError("Invalid file type. Only .xlsx files are supported.")
        
        # Check if file exists
        if not os.path.exists(full_excel_path):
            base_dir = os.path.dirname(os.path.abspath(__file__))
            raise FileNotFoundError(f"Excel file not found: {os.path.basename(full_excel_path)}\n"
                                  f"Expected location: {full_excel_path}\n"
                                  f"Make sure the file exists in: {base_dir}")
        
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(full_excel_path, data_only=True)
        
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")
        
        sheet = workbook[sheet_name]
        
        # Convert column letter to index (A = 1, B = 2, etc.)
        try:
            column_index = openpyxl.utils.column_index_from_string(column_letter)
        except ValueError:
            workbook.close()
            raise ValueError(f"Invalid column letter: {column_letter}")
        
        # Read all non-empty values in the column, excluding the header
        values = [
            str(sheet.cell(row=row_idx, column=column_index).value)
            for row_idx in range(2, sheet.max_row + 1)  # Start from row 2 to skip header
            if sheet.cell(row=row_idx, column=column_index).value is not None
        ]
        
        workbook.close()
        
        # Join values into a single comma-separated string
        output_string = ", ".join(values)
        return ([output_string],)  # Output as a list

NODE_CLASS_MAPPINGS = {"exLoadoutReadColumn": exLoadoutReadColumn}
NODE_DISPLAY_NAME_MAPPINGS = {"exLoadoutReadColumn": "exLoadout Read Column"}
