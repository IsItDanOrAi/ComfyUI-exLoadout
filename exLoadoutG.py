import openpyxl
import os

def get_excel_full_path_or_raise(base_folder, file_path):
    """
    Securely resolve Excel file paths within a designated directory.
    
    Args:
        base_folder: The base folder name (use "." for current directory)
        file_path: The requested file path
        
    Returns:
        str: The absolute path if valid
        
    Raises:
        ValueError: If the path is invalid or outside the allowed directory
    """
    # Get the directory where the script is located
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # If base_folder is ".", use the current directory, otherwise create subdirectory path
    if base_folder == ".":
        base_dir = current_dir
    else:
        base_dir = os.path.join(current_dir, base_folder)
    
    # Normalize the file path to prevent directory traversal
    normalized_file_path = os.path.normpath(file_path)
    
    # Check for directory traversal attempts
    if os.path.isabs(normalized_file_path) or normalized_file_path.startswith('..'):
        raise ValueError("Invalid file path. Absolute paths and parent directory references are not allowed.")
    
    # Construct the full path
    full_path = os.path.join(base_dir, normalized_file_path)
    
    # Resolve any remaining relative components
    resolved_path = os.path.abspath(full_path)
    
    # Ensure the resolved path is still within the base directory
    if not resolved_path.startswith(os.path.abspath(base_dir)):
        raise ValueError("Invalid file path. Path must be within the designated directory.")
    
    return resolved_path

# Hack: string type that is always equal in not equal comparisons
class AnyType(str):
    def __ne__(self, __value: object) -> bool:
        return False

ANY = AnyType("*")

class exLoadoutSeg2:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "exLoadoutList.xlsx"}),
                "sheet_name": ("STRING", {"default": "KSAMPLER"}),
                "row_number": ("INT", {"default": 1, "min": 1, "max": 1000, "step": 1}),
                "search_string": ("STRING", {"default": ""}),
            }
        }

    RETURN_TYPES = (ANY, ANY, ANY, ANY, ANY, ANY, "STRING")
    RETURN_NAMES = ("Column G", "Column H", "Column I", "Column J", "Column K", "Column L", "Outputs")
    OUTPUT_IS_LIST = (True, True, True, True, True, True, False)
    FUNCTION = "process_excel"
    CATEGORY = "exLoadout"
    DESCRIPTION = ("Reads values from columns G through L for a specified row number in an Excel spreadsheet. "
                   "Can also search for a string in Column A.")
    NAME = "exLoadoutSeg2 (List)"

    def process_excel(self, excel_path, sheet_name, row_number, search_string):
        # Secure path resolution for Excel file - look in current directory
        full_excel_path = get_excel_full_path_or_raise(".", excel_path)

        # Validate file extension
        if not full_excel_path.lower().endswith(".xlsx"):
            raise ValueError("Invalid file type. Only .xlsx files are supported.")

        # Check if file exists
        if not os.path.exists(full_excel_path):
            base_dir = os.path.dirname(os.path.abspath(__file__))
            raise FileNotFoundError(f"Excel file not found: {os.path.basename(full_excel_path)}\n"
                                  f"Expected location: {full_excel_path}\n"
                                  f"Make sure the file exists in: {base_dir}")

        workbook = openpyxl.load_workbook(full_excel_path)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")

        sheet = workbook[sheet_name]

        # Determine actual row based on search
        actual_row = row_number
        if search_string:
            found = False
            for row_idx in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=1).value
                if cell_value is not None and str(cell_value).strip() == search_string:
                    actual_row = row_idx
                    found = True
                    break
            if not found:
                workbook.close()
                raise ValueError(f"Search string '{search_string}' not found in Column A.")

        if actual_row < 1 or actual_row > sheet.max_row:
            workbook.close()
            raise ValueError(f"Row number {actual_row} is out of range. The sheet has {sheet.max_row} rows.")

        # Read columns G to L (7 to 12)
        row_data = [sheet.cell(row=actual_row, column=col_idx).value for col_idx in range(7, 13)]
        row_data = ['' if value is None else value for value in row_data]

        workbook.close()

        outputs_summary = (
            f"%G: {row_data[0]} %H: {row_data[1]} %I: {row_data[2]} "
            f"%J: {row_data[3]} %K: {row_data[4]} %L: {row_data[5]} %"
        )

        return ([row_data[0]], [row_data[1]], [row_data[2]],
                [row_data[3]], [row_data[4]], [row_data[5]], outputs_summary)

NODE_CLASS_MAPPINGS = {"exLoadoutSeg2": exLoadoutSeg2}
NODE_DISPLAY_NAME_MAPPINGS = {"exLoadoutSeg2": "exLoadout Seg2 (List)"}
