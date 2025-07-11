import openpyxl
import os
import tkinter as tk
from tkinter import ttk

def get_excel_full_path_or_raise(base_folder, file_path):
    """
    Securely resolve Excel file paths within a designated directory.
    
    Args:
        base_folder: The base folder name (use "." for current directory)
        file_path: The requested file path
        
    Returns:
        str: The absolute path if valid
        
    Raises:
        ValueError: If the path is invalid or outside the allowed directory
    """
    # Get the directory where the script is located
    current_dir = os.path.dirname(os.path.abspath(__file__))
    
    # If base_folder is ".", use the current directory, otherwise create subdirectory path
    if base_folder == ".":
        base_dir = current_dir
    else:
        base_dir = os.path.join(current_dir, base_folder)
    
    # Normalize the file path to prevent directory traversal
    normalized_file_path = os.path.normpath(file_path)
    
    # Check for directory traversal attempts
    if os.path.isabs(normalized_file_path) or normalized_file_path.startswith('..'):
        raise ValueError("Invalid file path. Absolute paths and parent directory references are not allowed.")
    
    # Construct the full path
    full_path = os.path.join(base_dir, normalized_file_path)
    
    # Resolve any remaining relative components
    resolved_path = os.path.abspath(full_path)
    
    # Ensure the resolved path is still within the base directory
    if not resolved_path.startswith(os.path.abspath(base_dir)):
        raise ValueError("Invalid file path. Path must be within the designated directory.")
    
    return resolved_path

class AnyType(str):
    def __ne__(self, __value: object) -> bool:
        return False

ANY = AnyType("*")

class exLoadoutEditCell:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "exLoadoutList.xlsx"}),
                "sheet_name": ("STRING", {"default": "Loadout_1"}),
                "row_number": ("INT", {"default": 1, "min": 1, "max": 10000}),
                "column_letter": (
                    ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"],
                    {"default": "A"}
                ),
                "new_value": ("STRING", {"default": ""}),
            },
        }

    RETURN_TYPES = (ANY,)
    RETURN_NAMES = ("output_row",)
    OUTPUT_IS_LIST = (True,)
    FUNCTION = "edit_excel_cell"
    CATEGORY = "exLoadout"
    DESCRIPTION = (
        "Edits a specific cell in an Excel spreadsheet and returns the entire row's values "
        "from columns A to L as a comma-separated string inside a list."
    )

    def edit_excel_cell(self, excel_path, sheet_name, row_number, column_letter, new_value):
        # ✅ Secure path resolution for Excel file - look in current directory
        full_excel_path = get_excel_full_path_or_raise(".", excel_path)

        # Validate file extension
        if not full_excel_path.lower().endswith(".xlsx"):
            raise ValueError("Invalid file type. Only .xlsx files are supported.")

        # Check if file exists
        if not os.path.exists(full_excel_path):
            base_dir = os.path.dirname(os.path.abspath(__file__))
            raise FileNotFoundError(f"Excel file not found: {os.path.basename(full_excel_path)}\n"
                                  f"Expected location: {full_excel_path}\n"
                                  f"Make sure the file exists in: {base_dir}")

        # Load the workbook
        workbook = openpyxl.load_workbook(full_excel_path)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file")

        sheet = workbook[sheet_name]
        column_index = openpyxl.utils.column_index_from_string(column_letter)

        if row_number < 1 or row_number > sheet.max_row:
            workbook.close()
            raise ValueError(f"Row {row_number} is out of range. The sheet has {sheet.max_row} rows.")

        if column_index < 1 or column_index > 12:
            workbook.close()
            raise ValueError(f"Column '{column_letter}' is out of the allowed range A-L.")

        # Edit the cell
        sheet.cell(row=row_number, column=column_index).value = new_value
        workbook.save(full_excel_path)

        # Retrieve updated row values from A–L
        row_values = []
        for col_idx in range(1, 13):
            cell_value = sheet.cell(row=row_number, column=col_idx).value
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_values.append(f"{col_letter}{row_number}: {str(cell_value)}")

        workbook.close()
        output_string = ", ".join(row_values)
        return ([output_string],)

    def create_edit_button(self):
        root = tk.Tk()
        root.title("Excel Editor")

        def on_edit():
            try:
                self.edit_excel_cell(
                    self.excel_path,
                    self.sheet_name,
                    self.row_number,
                    self.column_letter,
                    self.new_value
                )
                tk.messagebox.showinfo("Success", "Cell edited successfully.")
            except Exception as e:
                tk.messagebox.showerror("Error", str(e))

        edit_button = ttk.Button(root, text="Edit", command=on_edit)
        edit_button.pack(pady=20)
        root.mainloop()

NODE_CLASS_MAPPINGS = {"exLoadoutEditCell": exLoadoutEditCell}
NODE_DISPLAY_NAME_MAPPINGS = {"exLoadoutEditCell": "exLoadout Edit Cell"}
